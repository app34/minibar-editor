#!/usr/bin/env python3
"""Update the EXISTING minibar workbook in place.

Only cell values are written. Colours, merges, widths and formulas stay.
Usage:
  python update_inplace.py workbook.xlsx --day 1 --room 102 --item "MB Salted Peanuts" --qty 2
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def daily_sheet(wb, day: int):
    for name in wb.sheetnames:
        if name.strip().startswith(str(day)) and "August" in name:
            return wb[name]
    raise SystemExit(f"No daily sheet for day {day}")


def find_room_col(ws, room: str) -> int:
    for cell in ws[2]:
        if cell.value is None:
            continue
        if str(cell.value).strip() == str(room).strip():
            return cell.column
    raise SystemExit(f"Room {room} not found on {ws.title}")


def find_item_row(ws, item: str) -> int:
    needle = item.lower()
    for row in ws.iter_rows(min_row=3, max_col=2):
        name = row[1].value
        if name and needle in str(name).lower():
            return row[0].row
    raise SystemExit(f"Item {item!r} not found on {ws.title}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("workbook")
    p.add_argument("--day", type=int, required=True)
    p.add_argument("--room", required=True)
    p.add_argument("--item", required=True)
    p.add_argument("--qty", type=float, required=True)
    args = p.parse_args()

    path = Path(args.workbook)
    wb = load_workbook(path)
    ws = daily_sheet(wb, args.day)
    col = find_room_col(ws, args.room)
    row = find_item_row(ws, args.item)
    ws.cell(row=row, column=col).value = args.qty if args.qty else None
    wb.save(path)
    print(f"Updated {path.name}  {ws.title}  room {args.room}  row {row} col {col} = {args.qty}")


if __name__ == "__main__":
    main()
